import openpyxl
import os

USER_FILE = "users.xlsx"
JOURNAL_FILE = "journal_entries.xlsx"


def initialize_files():
    """Ensure the required Excel files exist."""
  
    # Initialize users file
    if not os.path.exists(USER_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["Name", "Password", "Index"])  # Header row
        ws.append(["admin", "admin123", 1])  # Default admin user
        wb.save(USER_FILE)

    # Initialize journal entries file
    if not os.path.exists(JOURNAL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "JournalEntries"
        ws.append(["Index", "Name", "Title", "Date", "Content"])  # Header row
        wb.save(JOURNAL_FILE)


def authenticate(username, password):
    """Check if username and password match any existing user and verify their role using index."""
    wb = openpyxl.load_workbook(USER_FILE)
    ws = wb["Users"]

    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] == username and row[1] == password:
            role = "admin" if row[2] == 1 else "user"  # Check index for admin
            return role
    return None


def signup():
    """Sign up a new user."""
    username = input("Enter username: ")
    password = input("Enter password: ")

    wb = openpyxl.load_workbook(USER_FILE)
    ws = wb["Users"]

    new_index = ws.max_row  # Assign new index based on the last row
    ws.append([username, password, new_index])
    wb.save(USER_FILE)
    print("User signed up successfully!")


def add_journal_entry(username):
    """Add a new journal entry."""
    title = input("Enter journal title: ")
    date = input("Enter journal date (YYYY-MM-DD): ")
    content = input("Enter journal content: ")

    wb = openpyxl.load_workbook(JOURNAL_FILE)
    ws = wb["JournalEntries"]

    new_index = ws.max_row  # Assign new index based on the last row
    ws.append([new_index, username, title, date, content])
    wb.save(JOURNAL_FILE)
    print("Journal entry added successfully!")


def delete_journal_entry(username):
    """Delete a journal entry."""
    index = input("Enter the index of the journal entry to delete: ")
    wb = openpyxl.load_workbook(JOURNAL_FILE)
    ws = wb["JournalEntries"]

    found = False
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if row[0].value == int(index) and row[1].value == username:
            ws.delete_rows(row[0].row)
            found = True
            break

    if found:
        wb.save(JOURNAL_FILE)
        print("Journal entry deleted successfully.")
    else:
        print("No matching entry found.")


def edit_journal_entry(username):
    """Edit an existing journal entry."""
    index = input("Enter the index of the journal entry to edit: ")
    wb = openpyxl.load_workbook(JOURNAL_FILE)
    ws = wb["JournalEntries"]

    found = False
    for row in ws.iter_rows(min_row=2):  # Skip header row
        if row[0].value == int(index) and row[1].value == username:
            print("Current entry:", {cell.column_letter: cell.value for cell in row})
            row[2].value = input("Enter new title: ")
            row[3].value = input("Enter new date (YYYY-MM-DD): ")
            row[4].value = input("Enter new content: ")
            found = True
            break

    if found:
        wb.save(JOURNAL_FILE)
        print("Journal entry updated successfully.")
    else:
        print("No matching entry found.")


def search_journal_entries(username):
    """Search for journal entries by title or date."""
    query = input("Enter a title or date to search for: ")
    wb = openpyxl.load_workbook(JOURNAL_FILE)
    ws = wb["JournalEntries"]

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[1] == username and (query in row[2] or query in row[3]):
            results.append(row)

    if results:
        print("Search results:")
        for result in results:
            print(result)
    else:
        print("No matching entries found.")


def main_menu():
    """Main menu to log in or sign up."""
    while True:
        print("\nMain Menu")
        print("1. Login")
        print("2. Sign Up")
        print("3. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            username = input("Enter username: ")
            password = input("Enter password: ")
            role = authenticate(username, password)

            if role:
                print(f"Welcome {username}! Role: {role}")
                if role == "admin":
                    admin_actions(username)
                else:
                    user_actions(username)
            else:
                print("Invalid credentials.")
        elif choice == "2":
            signup()
        elif choice == "3":
            print("Goodbye!")
            break
        else:
            print("Invalid option. Try again.")


def admin_actions(username):
    """Admin-specific actions."""
    while True:
        print("\nAdmin Menu")
        print("1. Add Journal Entry")
        print("2. Delete Journal Entry")
        print("3. Edit Journal Entry")
        print("4. Search Journal Entries")
        print("5. Logout")
        choice = input("Choose an option: ")

        if choice == "1":
            add_journal_entry(username)
        elif choice == "2":
            delete_journal_entry(username)
        elif choice == "3":
            edit_journal_entry(username)
        elif choice == "4":
            search_journal_entries(username)
        elif choice == "5":
            break
        else:
            print("Invalid option. Try again.")


def user_actions(username):
    """User-specific actions."""
    while True:
        print("\nUser Menu")
        print("1. Add Journal Entry")
        print("2. Search Journal Entries")
        print("3. Logout")
        choice = input("Choose an option: ")

        if choice == "1":
            add_journal_entry(username)
        elif choice == "2":
            search_journal_entries(username)
        elif choice == "3":
            break
        else:
            print("Invalid option. Try again.")


# Initialize files and run the program
initialize_files()
main_menu()
